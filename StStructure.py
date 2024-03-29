import streamlit as st
import pandas as pd
import math
import sympy as sp
from openpyxl import load_workbook


st.title('App Cálculo estrutural 2D')

def Prop():
    Propriedades = []
    MI = []
    ME =[]
    AR = []
    filename = 'Valores.xlsx'
    planilha = load_workbook(filename)
    i=1
    while (AbaProp.cell(i,1).value) != None:
        MI.append(AbaProp.cell(i,1).value)
        ME.append(AbaProp.cell(i,2).value)
        AR.append(AbaProp.cell(i,3).value)
        i +=1
    Propriedades.append(MI)
    Propriedades.append(ME)
    Propriedades.append(AR)
    return Propriedades
    
def Nosf():
    Nos = []
    x = []
    y = []
    filename = 'Valores.xlsx'
    planilha = load_workbook(filename)
    i=1
    while (AbaNos.cell(i,1).value) != None:
        No = []
        No.append(AbaNos.cell(i,1).value)
        No.append(AbaNos.cell(i,2).value)
        Nos.append(No)
        i +=1
    return Nos


def barras():
    barras = []
    filename = 'Valores.xlsx'
    planilha = load_workbook(filename)
    i=1
    while (AbaBarras.cell(i,1).value) != None:
        barra = []
        barra.append(AbaBarras.cell(i,1).value)
        barra.append(AbaBarras.cell(i,2).value)
        barra.append(AbaBarras.cell(i,3).value)
        barra.append(AbaBarras.cell(i,4).value)
        barra.append(AbaBarras.cell(i,5).value)
        barras.append(barra)
        i +=1
    return barras

def Forcas():
    forcas = []
    filename = 'Valores.xlsx'
    planilha = load_workbook(filename)
    i=1
    while (AbaForca.cell(i,1).value) != None:
        forca = []
        forca.append(AbaForca.cell(i,1).value)
        forca.append(AbaForca.cell(i,2).value)
        forca.append(AbaForca.cell(i,3).value)
        forca.append(AbaForca.cell(i,4).value)
        forcas.append(forca)
        i +=1
    return forcas

def Apoios():
    apoios = []
    filename = 'Valores.xlsx'
    planilha = load_workbook(filename)
    i=1
    while (AbaApoio.cell(i,1).value) != None:
        apoio = []
        apoio.append(AbaApoio.cell(i,1).value)
        apoio.append(AbaApoio.cell(i,2).value)
        apoio.append(AbaApoio.cell(i,3).value)
        apoio.append(AbaApoio.cell(i,4).value)
        apoios.append(apoio)
        i +=1
    return apoios

filename = 'Valores.xlsx'
planilha = load_workbook(filename)

AbaProp = planilha['AbaProp']
AbaNos = planilha['AbaNos']
AbaBarras = planilha['AbaBarras']
AbaForca = planilha['AbaForca']
AbaApoio = planilha['AbaApoio']

st.sidebar.header('Propriedades')
I = st.sidebar.number_input('Momento de Inércia:')
E = st.sidebar.number_input('Módulo de Elasticidade:')          
A = st.sidebar.number_input('Área:')

if st.sidebar.button('Confirmar'):
    i=1
    while (AbaProp.cell(i,1).value) != None:
        i +=1
    AbaProp.cell(i,1).value = I
    AbaProp.cell(i,2).value = E
    AbaProp.cell(i,3).value = A
    planilha.save(filename)

st.sidebar.header('Nós')
x = st.sidebar.number_input('Insira coordenada x')
y = st.sidebar.number_input('Insira coordenada y')

if st.sidebar.button('Confirmar Nó'):
    i=1
    while (AbaNos.cell(i,1).value) != None:
        i +=1
    AbaNos.cell(i,1).value = x
    AbaNos.cell(i,2).value = y
    planilha.save(filename)


nos = Nosf()
Propriedades = Prop()
nprop = len(Propriedades[0])
if (nprop == 0):
    nprop=1


if st.sidebar.button('Mostrar Propriedades'):
    st.write(Propriedades)
if st.sidebar.button('Mostrar Nos'):
    st.write(nos)

st.sidebar.header('Barras')
no1 = st.sidebar.number_input('Insira o nó1',value=0)
no2 = st.sidebar.number_input('Insira o nó2',value=0)
prop = st.sidebar.slider('Selecione um padrão de barra',0,nprop,0)

if st.sidebar.button('Confirmar barra'):
    i=1
    while (AbaBarras.cell(i,1).value) != None:
        i +=1
    AbaBarras.cell(i,1).value = no1
    AbaBarras.cell(i,2).value = no2
    AbaBarras.cell(i,3).value = Propriedades[0][prop]
    AbaBarras.cell(i,4).value = Propriedades[1][prop]
    AbaBarras.cell(i,5).value = Propriedades[2][prop]
    planilha.save(filename)
    
barras = barras()

if st.sidebar.button('Mostrar Barras'):
    st.write(barras)

st.sidebar.header('Forças')

no = st.sidebar.number_input('Insira o nó',value=0)
Fx = st.sidebar.number_input('Insira a decomposição x da Força')
Fy = st.sidebar.number_input('Insira a decomposição y da Força')
M = st.sidebar.number_input('Insira o momento')

if st.sidebar.button('Confirmar Força'):
    i=1
    while (AbaForca.cell(i,1).value) != None:
        i +=1
    AbaForca.cell(i,1).value = no
    AbaForca.cell(i,2).value = Fx
    AbaForca.cell(i,3).value = Fy
    AbaForca.cell(i,4).value = M
    planilha.save(filename)

forcas = Forcas()

if st.sidebar.button('Mostrar Forças'):
    st.write(forcas)

st.sidebar.header('Apoios')

noap = st.sidebar.number_input('Insira o nó do apoio',value=0)
restricaox = st.sidebar.checkbox ('O movimento é restrito em x?')
restricaoy = st.sidebar.checkbox ('O movimento é restrito em y?')
restricaoz = st.sidebar.checkbox ('O movimento é restrito em z?')

if st.sidebar.button('Confirmar Apoio'):
    i=1
    while (AbaApoio.cell(i,1).value) != None:
        i +=1
    AbaApoio.cell(i,1).value = noap
    if restricaox:
        AbaApoio.cell(i,2).value = True
    else:
        AbaApoio.cell(i,2).value = False
    if restricaoy:
        AbaApoio.cell(i,3).value = True
    else:
        AbaApoio.cell(i,3).value = False
    if restricaoz:
        AbaApoio.cell(i,4).value = True
    else:
        AbaApoio.cell(i,4).value = False
    planilha.save(filename)

apoios = Apoios()

if st.sidebar.button('Mostrar Apoios'):
    st.write(apoios)

if st.sidebar.button('Limpar arquivo'):
    AbaProp.delete_cols(3)
    AbaProp.delete_cols(2)
    AbaProp.delete_cols(1)
    AbaNos.delete_cols(2)
    AbaNos.delete_cols(1)
    AbaBarras.delete_cols(5)
    AbaBarras.delete_cols(4)
    AbaBarras.delete_cols(3)
    AbaBarras.delete_cols(2)
    AbaBarras.delete_cols(1)
    AbaForca.delete_cols(4)
    AbaForca.delete_cols(3)
    AbaForca.delete_cols(2)
    AbaForca.delete_cols(1)
    AbaApoio.delete_cols(4)
    AbaApoio.delete_cols(3)
    AbaApoio.delete_cols(2)
    AbaApoio.delete_cols(1)
    planilha.save(filename)


QTDE_BARRAS = len(barras)

# Cálculo do comprimento e ângulo de cada barra
for barra in barras:
    no1 = nos[barra[0]]
    no2 = nos[barra[1]]
    barra.append(((no2[0] - no1[0]) ** 2 + (no2[1] - no1[1]) ** 2) ** 0.5)
    barra.append(math.atan2(no2[1] - no1[1], no2[0] - no1[0]))

# Quantidade de nós e graus de liberdade
QTDE_NOS = len(nos)
GDL = 3 * QTDE_NOS

# Definição dos símbolos
A, E, L, I, theta = sp.symbols(["A", "E", "L", "I", "theta"])

# Definição das variáveis para a matriz de rigidez
COS = sp.cos(theta)
SIN = sp.sin(theta)
T = sp.Matrix(
    [
        [COS, -SIN, 0, 0, 0, 0],
        [SIN, COS, 0, 0, 0, 0],
        [0, 0, 1, 0, 0, 0],
        [0, 0, 0, COS, -SIN, 0],
        [0, 0, 0, SIN, COS, 0],
        [0, 0, 0, 0, 0, 1],
    ]
).T

Ke_ = sp.Matrix(
    [
        [E * A / L, 0, 0, -E * A / L, 0, 0],
        [
            0,
            12 * E * I / L**3,
            6 * E * I / L**2,
            0,
            -12 * E * I / L**3,
            6 * E * I / L**2,
        ],
        [0, 6 * E * I / L**2, 4 * E * I / L, 0, -6 * E * I / L**2, 2 * E * I / L],
        [-E * A / L, 0, 0, E * A / L, 0, 0],
        [
            0,
            -12 * E * I / L**3,
            -6 * E * I / L**2,
            0,
            12 * E * I / L**3,
            -6 * E * I / L**2,
        ],
        [0, 6 * E * I / L**2, 2 * E * I / L, 0, -6 * E * I / L**2, 4 * E * I / L],
    ]
)

Ke = T.T * Ke_ * T

# Montagem da matriz de rigidez
lista_Ke = []
for i in range(QTDE_BARRAS):
    lista_Ke.append(
        Ke.subs(
            [
                (I, barras[i][2]),
                (E, barras[i][3]),
                (A, barras[i][4]),
                (L, barras[i][5]),
                (theta, barras[i][6]),
            ]
        )
    )

# Montagem da matriz de rigidez global
K = sp.zeros(GDL, GDL)
for i in range(QTDE_BARRAS):
    no1 = barras[i][0]
    no2 = barras[i][1]
    indices = [3 * no1, 3 * no1 + 1, 3 * no1 + 2, 3 * no2, 3 * no2 + 1, 3 * no2 + 2]
    for j in range(6):
        for k in range(6):
            K[indices[j], indices[k]] += lista_Ke[i][j, k]

# Montagem do vetor de forças
S = [0] * GDL
for i in range(len(forcas)):
    no = forcas[i][0]
    ForcaX = forcas[i][1]
    ForcaY = forcas[i][2]
    momento = forcas[i][3]
    if type(ForcaX) == str:
        S[3* no] = sp.symbols(["S"+str(3*no+1)])
    else:
        S[3 * no] += ForcaX
    if type(ForcaY) ==str:
        S[3 * no + 1] = sp.symbols(["S"+str(3*no+2)])
    else:
        S[3 * no + 1] += ForcaY
    if type(momento) == str:
        S[3 * no + 2] = sp.symbols(["S"+str(3*no+3)])
    else: 
        S[3 * no + 2] += momento
S = sp.Matrix(S)

# Montagem do vetor de deslocamentos
q = sp.symbols(["q" + str(i) for i in range(1, GDL + 1)])
for i in range(len(apoios)):
    no = apoios[i][0]
    if apoios[i][1]:
        q[3 * no] = 0
        S[3 * no] = sp.symbols("S" + str(3 * no + 1))
    if apoios[i][2]:
        q[3 * no + 1] = 0
        S[3 * no + 1] = sp.symbols("S" + str(3 * no + 2))
    if apoios[i][3]:
        q[3 * no + 2] = 0
        S[3 * no + 2] = sp.symbols("S" + str(3 * no + 3))
q = sp.Matrix(q)

# Separando variáveis a serem resolvidas
variaveis_sistema = [i for i in q if isinstance(i, sp.Symbol)] + [
    i for i in S if isinstance(i, sp.Symbol)
]

# Resolução do sistema
sistema = K * q - S
resolucao = sp.solve(sistema, variaveis_sistema)


lista_Ke_ = []
for i in range(QTDE_BARRAS):
    lista_Ke_.append(
        Ke_.subs(
            [
                (I, barras[i][2]),
                (E, barras[i][3]),
                (A, barras[i][4]),
                (L, barras[i][5]),
            ]
        )
    )

def Esforco_barra(numero_da_barra):
    n = numero_da_barra
    nosapoios=[]
    for i in range(len(apoios)):
        nosapoios.append(apoios[i][0])

    qnapoios = []
    for i in range(len(apoios)):
        noq = apoios[i][0]
        if apoios[i][1]:
            qnapoios.append('q' + str(3*noq+1))
        if apoios[i][2]:
            qnapoios.append('q'+str(3*noq+2))
        if apoios[i][3]:
            qnapoios.append('q'+str(3*noq+3))
    
     
    nob1 = barras[n][0]
    nob2 = barras[n][1]
    thetab = barras[n][6]
    cb = sp.cos(thetab)
    sb = sp.sin(thetab)
    COS = cb
    SIN = sb

    #Encontrando o valor das forças aplicadas a cada nó
    Snforcas = []
    S1b = 0
    S2b = 0
    S3b = 0
    S4b = 0
    S5b = 0
    S6b = 0
    for i in range(len(forcas)):
        nof = forcas[i][0]
        ForcaX = forcas[i][1]
        ForcaY = forcas[i][2]
        momento = forcas[i][3]
        if (nob1 == nof):
            S1b = ForcaX
            S2b = ForcaY
            S3b = momento
        if (nob2 == nof):
            S4b = ForcaX
            S5b = ForcaY
            S6b = momento
    
    Sb = sp.Matrix([S1b,S2b,S3b,S4b,S5b,S6b])
    #definindo forças locais aplicadas na barra
    Tb =  sp.Matrix(
    [
        [COS, -SIN, 0, 0, 0, 0],
        [SIN, COS, 0, 0, 0, 0],
        [0, 0, 1, 0, 0, 0],
        [0, 0, 0, COS, -SIN, 0],
        [0, 0, 0, SIN, COS, 0],
        [0, 0, 0, 0, 0, 1],
    ])
    Sb_ = Tb*Sb

    #Definindo a nomenclatura de cada deslocamento q para os nós da barra escolhida
    U1 = ("q" + str(3*nob1+1))
    U2 = ("q" + str(3*nob1+2))
    U3 = ("q" + str(3*nob1+3))
    U4 = ("q" + str(3*nob1+4))
    U5 = ("q" + str(3*nob1+5))
    U6 = ("q" + str(3*nob1+6))

    if (nob1 in nosapoios):
        if (U1 in qnapoios):
            U1 = 0
        if (U2 in qnapoios):
            U2 = 0
        if (U3 in qnapoios):
            U3 = 0

    if (nob2 in nosapoios):
        if (U4 in qnapoios):
            U4 = 0
        if (U5 in qnapoios):
            U5 = 0
        if (U6 in qnapoios):
            U6 = 0


    if (U1 == 0):
        qb1 = 0
    else:
        qb1 = resolucao[sp.symbols(U1)]

    if (U2 == 0):
        qb2 = 0
    else:
        qb2 = resolucao[sp.symbols(U2)]

    if (U3 == 0):
        qb3 = 0
    else:
        qb3 = resolucao[sp.symbols(U3)]

    if (U4 == 0):
        qb4 = 0
    else:
        qb4 = resolucao[sp.symbols(U4)]

    if (U5 == 0):
        qb5 = 0
    else:
        qb5 = resolucao[sp.symbols(U5)]

    if (U6 == 0):
        qb6 = 0
    else:
        qb6 = resolucao[sp.symbols(U6)]

    qb = sp.Matrix([qb1,qb2,qb3,qb4,qb5,qb6])  
    Kl = lista_Ke_[n]
    COS = cb
    SIN = sb
    T_ = sp.Matrix(
        [
            [COS, -SIN, 0, 0, 0, 0],
            [SIN, COS, 0, 0, 0, 0],
            [0, 0, 1, 0, 0, 0],
            [0, 0, 0, COS, -SIN, 0],
            [0, 0, 0, SIN, COS, 0],
            [0, 0, 0, 0, 0, 1],
        ]
    ).T
    u1,u2,u3,u4,u5,u6 = 'u1','u2','u3','u4','u5','u6'
    ub = sp.Matrix([u1,u2,u3,u4,u5,u6])
    sistema2 = T_*qb-ub
    ub_local = sp.solve(sistema2,[u1,u2,u3,u4,u5,u6])
    u_1 = ub_local[sp.symbols('u1')]
    u_2 = ub_local[sp.symbols('u2')]
    u_3 = ub_local[sp.symbols('u3')]
    u_4 = ub_local[sp.symbols('u4')]
    u_5 = ub_local[sp.symbols('u5')]
    u_6 = ub_local[sp.symbols('u6')]
    ub_ = sp.Matrix([u_1,u_2,u_3,u_4,u_5,u_6])
    S1b_,S2b_,S3b_,S4b_,S5b_,S6b_ = 'S1b_','S2b_','S3b_','S4b_','S5b_','S6b_'
    S_ = sp.Matrix([S1b_,S2b_,S3b_,S4b_,S5b_,S6b_])
    sistema3 = Kl*ub_-S_+Sb_
    Esforcos = sp.solve(sistema3,[S1b_,S2b_,S3b_,S4b_,S5b_,S6b_])

    S1_ = Esforcos[sp.symbols(S1b_)]
    S2_ = -Esforcos[sp.symbols(S2b_)]
    S3_ = Esforcos[sp.symbols(S3b_)]
    S4_ = Esforcos[sp.symbols(S4b_)]
    S5_ = Esforcos[sp.symbols(S5b_)]
    S6_ = -Esforcos[sp.symbols(S6b_)]
    q1_ =u_1
    q2_ =u_2
    q3_ =u_3
    q4_ =u_4
    q5_ =u_5
    q6_ =u_6
    results = sp.Matrix([S1_,S2_,S3_,S4_,S5_,S6_,q1_,q2_,q3_,q4_,q5_,q6_])
    return results

def Momento_Fletor(E_,I_,L_,M_0,M_3,q_2,q_3,q_5,q_6,xm_):
    
    x = sp.symbols('x')
    E = E_
    I = I_
    L = L_


    a = sp.Matrix([[0,0,2,6*x]])
    b = sp.Matrix([
        [1,0,0,0],
        [0,1,0,0],
        [-3/(L**2),-2/L,3/(L**2),-1/L],
        [2/(L**3),1/(L**2),-2/(L**3),1/(L**2)]
    ])

    N2x = a*b

    x0 = 0
    x3 = L
    x1 = (L/2)*(1-1/(3**0.5))
    x2 = L+((L/2)*(1/(3**0.5)))


    q2=q_2
    q3=q_3
    q5=q_5
    q6=q_6
    d_nodais = sp.Matrix([q2,q3,q5,q6])

    d_interno = N2x*d_nodais

    y1 = d_interno.subs(x,x1)
    y2 = d_interno.subs(x,x2)

    M0 = M_0
    M3 = M_3

    M1 = y1*E*I
    M2 = y2*E*I

    fx0 = M0
    fx1 = M1[0]
    fx2 = M2[0]
    fx3 = M3

    fx0x1 = (fx1-fx0)/(x1-x0)
    fx1x3 = (fx3-fx1)/(x3-x1)
    fx0x1x3 = (fx1x3-fx0x1)/(x3-x0)

    d0 = fx0
    d1 = fx0x1
    d2 = fx0x1x3

    eq = d0 + d1*(x-x0) + d2*(x-x0)*(x-x1)

    xm = xm_
    a = eq.subs(x,xm)

    return a

if st.button('Calcular'): 
    resultado = str(resolucao)
    st.header('A resolução do sistema é:')
    st.write(resultado)

n_barra = 0
n_barra = st.number_input('Insira uma barra para calcular os esforços',0)
x_barra = 5
#x_barra = st.number_input('Escolha um ponto da barra para calcular o momento',0.00)
x_barra = st.slider('Escolha um ponto x da barra para calcular o momento', 0.00, barras[n_barra][5], 0.00)


if st.button('Esforços Barra'):
    Esforco_na_barra = Esforco_barra(n_barra)
    st.header('Os esforços na barra são:')
    st.subheader('Axial Nó 1')
    st.write (Esforco_na_barra[0])
    st.subheader('Axial Nó 2')
    st.write (Esforco_na_barra[3])
    st.subheader('Cortante Nó 1')
    st.write (Esforco_na_barra[1])
    st.subheader('Cortante Nó 2')
    st.write (Esforco_na_barra[4])
    st.subheader('Momento Nó 1')
    st.write (Esforco_na_barra[2])
    st.subheader('Momento Nó 2')
    st.write (Esforco_na_barra[5])
    Ib = barras[n_barra][2]
    Eb = barras[n_barra][3]
    Lb = barras[n_barra][5]
    Momento_em_x = Momento_Fletor(Eb,Ib,Lb,Esforco_na_barra[2],Esforco_na_barra[5],Esforco_na_barra[7],Esforco_na_barra[8],Esforco_na_barra[10],Esforco_na_barra[11],x_barra)
    st.subheader('O momento em para o valor de x informado é')
    st.write(Momento_em_x)














    


